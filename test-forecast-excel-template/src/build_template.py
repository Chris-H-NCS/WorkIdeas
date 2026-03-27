from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1F4E78")


def style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, min_width=12, max_width=42):
    for col_idx in range(1, ws.max_column + 1):
        col = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col].width = max(min_width, min(max_width, max_len + 2))


def freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_workbook(output_path: Path):
    wb = Workbook()

    # Instructions
    ws_i = wb.active
    ws_i.title = "Instructions"
    ws_i["A1"] = "Test Design & Execution Forecast Template"
    ws_i["A1"].font = TITLE_FONT
    ws_i["A3"] = "How to use"
    ws_i["A3"].font = Font(bold=True)
    ws_i["A4"] = "1) Fill assumptions in Inputs sheet"
    ws_i["A5"] = "2) Add planned test cases in Test_Plan"
    ws_i["A6"] = "3) Track execution in Execution_Tracking"
    ws_i["A7"] = "4) Review KPIs and forecast in Forecast_Summary"
    ws_i["A9"] = "Tip: Keep Test IDs consistent across Test_Plan and Execution_Tracking."
    ws_i["A9"].font = Font(italic=True, color="555555")
    ws_i.column_dimensions["A"].width = 90

    # Inputs
    ws_in = wb.create_sheet("Inputs")
    ws_in.append(["Input", "Value", "Notes"])
    rows = [
        ("Planned test cases", 120, "Total number of test cases expected"),
        ("Avg design hours per test", 1.5, "Average effort to design one test case"),
        ("Avg execution hours per test", 0.75, "Average effort to execute one test case"),
        ("No. of test designers", 2, "People available for design"),
        ("Design hours per designer/day", 6, "Net productive hours per day"),
        ("No. of test executors", 3, "People available for execution"),
        ("Execution hours per executor/day", 6, "Net productive hours per day"),
        ("Re-test uplift %", 0.15, "Expected re-test effort as a fraction"),
        ("Contingency %", 0.1, "Extra allowance for uncertainty"),
        ("Total design hours", "=B2*B3*(1+B10)", "Calculated"),
        ("Total execution hours", "=B2*B4*(1+B9+B10)", "Calculated"),
        ("Forecast design days", "=IF(B5*B6=0,0,B11/(B5*B6))", "Calculated"),
        ("Forecast execution days", "=IF(B7*B8=0,0,B12/(B7*B8))", "Calculated"),
    ]
    for row in rows:
        ws_in.append(row)
    style_header(ws_in)
    ws_in["A1"].font = HEADER_FONT
    ws_in["B1"].font = HEADER_FONT
    ws_in["C1"].font = HEADER_FONT
    for r in range(9, 14):
        ws_in[f"B{r+1}"].number_format = "0.00%"
    for r in [11, 12, 13, 14]:
        ws_in[f"B{r}"].number_format = "0.00"
    freeze_and_filter(ws_in)

    # Test Plan
    ws_tp = wb.create_sheet("Test_Plan")
    ws_tp.append([
        "Test ID",
        "Feature",
        "Test Scenario",
        "Priority",
        "Test Type",
        "Complexity",
        "Design Owner",
        "Estimated Design Hours",
        "Estimated Execution Hours",
        "Design Status",
    ])
    sample_plan = [
        ("TC-001", "Login", "Valid login with MFA", "High", "Functional", "Medium", "Analyst A", 1.5, 0.75, "Designed"),
        ("TC-002", "Login", "Invalid password lockout", "High", "Functional", "Low", "Analyst B", 1.0, 0.5, "In Design"),
        ("TC-003", "Checkout", "Promo code with tax", "Medium", "Integration", "High", "Analyst A", 2.0, 1.0, "Not Started"),
    ]
    for row in sample_plan:
        ws_tp.append(row)
    style_header(ws_tp)
    freeze_and_filter(ws_tp)

    # Execution Tracking
    ws_et = wb.create_sheet("Execution_Tracking")
    ws_et.append([
        "Test ID",
        "Feature",
        "Execution Status",
        "Planned Exec Date",
        "Actual Exec Date",
        "Defect ID",
        "Defect Severity",
        "Re-test Required",
        "Execution Owner",
        "Notes",
    ])
    sample_exec = [
        ("TC-001", "Login", "Pass", "2026-03-30", "2026-03-30", "", "", "No", "Tester A", ""),
        ("TC-002", "Login", "Fail", "2026-03-30", "2026-03-31", "BUG-1201", "High", "Yes", "Tester B", "Account lock threshold mismatch"),
        ("TC-003", "Checkout", "Not Run", "2026-04-01", "", "", "", "No", "Tester C", "Blocked by env refresh"),
    ]
    for row in sample_exec:
        ws_et.append(row)
    style_header(ws_et)
    freeze_and_filter(ws_et)

    # Forecast Summary
    ws_fs = wb.create_sheet("Forecast_Summary")
    ws_fs["A1"] = "Forecast Summary"
    ws_fs["A1"].font = TITLE_FONT

    summary_rows = [
        ("Total Planned Test Cases", "=Inputs!B2"),
        ("Designed Test Cases", "=COUNTIF(Test_Plan!J2:J1000,\"Designed\")"),
        ("Design Completion %", "=IF(B2=0,0,B3/B2)"),
        ("Executed Test Cases", "=COUNTIFS(Execution_Tracking!C2:C1000,\"<>Not Run\",Execution_Tracking!C2:C1000,\"<>\")"),
        ("Execution Completion %", "=IF(B2=0,0,B5/B2)"),
        ("Passed", "=COUNTIF(Execution_Tracking!C2:C1000,\"Pass\")"),
        ("Failed", "=COUNTIF(Execution_Tracking!C2:C1000,\"Fail\")"),
        ("Blocked", "=COUNTIF(Execution_Tracking!C2:C1000,\"Blocked\")"),
        ("Pass Rate %", "=IF(B5=0,0,B6/B5)"),
        ("Open Defect Count", "=COUNTIF(Execution_Tracking!F2:F1000,\"<>\")"),
        ("Forecast Design Days", "=Inputs!B13"),
        ("Forecast Execution Days", "=Inputs!B14"),
    ]

    ws_fs.append(["Metric", "Value"])
    for metric, formula in summary_rows:
        ws_fs.append([metric, formula])

    style_header(ws_fs, row=2)
    ws_fs["A2"].fill = HEADER_FILL
    ws_fs["B2"].fill = HEADER_FILL
    ws_fs["A2"].font = HEADER_FONT
    ws_fs["B2"].font = HEADER_FONT
    ws_fs["B4"].number_format = "0.00%"
    ws_fs["B6"].number_format = "0.00%"
    ws_fs["B10"].number_format = "0.00%"
    ws_fs["B12"].number_format = "0.00"
    ws_fs["B13"].number_format = "0.00"

    for ws in [ws_in, ws_tp, ws_et, ws_fs]:
        auto_width(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    project_root = Path(__file__).resolve().parents[1]
    out_file = project_root / "output" / "Test_Design_Execution_Forecast_Template.xlsx"
    build_workbook(out_file)
    print(f"Template created: {out_file}")
